import random
 
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
 
# 定义合并函数
def Merge_cells(result_file, sheet_name, row_num):
    wb = openpyxl.load_workbook(result_file)
    ws = wb[sheet_name]
 
    # 验证参数
    if row_num < 1:
        raise ValueError("col_num must be a positive integer.")
 
    # 获取指定行数据
    row_list = [cell[0].value for cell in ws.iter_cols(min_col=1, min_row=row_num, max_row=row_num)]
    print(row_list)
    if not row_list:
        return  # 列数据为空，无需合并
 
    # 合并单元格的逻辑
    merge_ranges = []
    count = 0;
    for i in range(1, len(row_list)):
        if row_list[i] != row_list[i - 1]:
            merge_ranges.append([i - count, i - 1])
            count = 0
        else:
            count = count + 1
 
    if len(merge_ranges) == 0 and count > 0:
        merge_ranges.append([0, len(row_list) - 1])
    else:
        merge_ranges.append([len(row_list) - count, len(row_list) - 1])
 
    # 合并单元格和设置对齐
    for start, end in merge_ranges:
        merge_range = f"{get_column_letter(1 if start == 0 else start)}{row_num}:{get_column_letter(end + 1)}{row_num}"
        print(merge_range)
        ws.merge_cells(merge_range)
        for row in ws[merge_range]:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
 
    wb.save(result_file)
 
 
if __name__ == '__main__':
    # 模拟三级标题
    data = [{'name': '湖北', 'city': [{'name': '武汉'}, {'name': '宜昌'}, {'name': '襄阳'}]},
            {'name': '湖南', 'city': [{'name': '长沙'}, {'name': '株洲'}, {'name': '湘潭'}]},
            {'name': '江苏', 'city': [{'name': '杭州'}, {'name': '苏州'}]}]
 
    title1 = []
    title2 = []
    title3 = []
    for i, datum in enumerate(data):
        city = list(datum['city'])
        for e in city:
            title1.append('中国')
            title2.append(datum['name'])
            title3.append(e['name'])
            
    rows = [title2, title3]
    
    for i in range(100):
        row = []
        for j in range(len(title2)):
            row.append(random.randint(1, 10))
        rows.append(row)
 
    df = pd.DataFrame(rows)  # 创建DataFrame 设置2,3级标题
    df.columns = title1  # 设置1级标题
   
    df.to_excel('text.xlsx', index=False)  # 存表，去除原始索引列（0,1,2...）
 
    Merge_cells('text.xlsx', 'Sheet1', 1)  # 合并一级标题
    Merge_cells('text.xlsx', 'Sheet1', 2)  # 合并第二行标题