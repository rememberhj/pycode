from openpyxl import load_workbook
 
# 加载刚才用pandas保存的Excel文件
wb = load_workbook("test2.xlsx")
ws = wb.active
 
# 真正地合并单元格（如果需要的话）并添加数据到左上角的单元格中（如果之前是空的）
ws.merge_cells('A1:C1')  # 假设你想合并第一行的A到C列的单元格们
ws['A1'] = "这是合并后的单元格"  # 确保左上角的单元格有数据，以防之前是空的字符串或NaN
 
# 保存修改后的文件
wb.save("test2.xlsx")